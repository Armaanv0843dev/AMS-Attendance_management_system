"""
services/excel.py — Excel report generation using OpenPyXL.
Generates a formatted .xlsx attendance report.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from io import BytesIO
from datetime import datetime


# Color palette (hex without #)
COLOR_PRIMARY = "4F46E5"
COLOR_SUCCESS = "10B981"
COLOR_WARNING = "F59E0B"
COLOR_DANGER = "EF4444"
COLOR_DARK = "1E1E1E"
COLOR_LIGHT = "F3F4F6"
COLOR_WHITE = "FFFFFF"
COLOR_HEADER_BG = "1E1E2E"


def thin_border():
    side = Side(style="thin", color="D1D5DB")
    return Border(left=side, right=side, top=side, bottom=side)


def build_excel_report(predictions: list, attendance: list = None) -> bytes:
    """
    Build an Excel attendance report.

    Args:
        predictions: List of prediction dicts.
        attendance:  Optional raw attendance list for the raw sheet.

    Returns:
        Excel file bytes.
    """
    wb = Workbook()

    # ===== Sheet 1: Summary =====
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary.merge_cells("A1:K1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Attendance Management System — Report"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=COLOR_WHITE)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 32

    # Generated date
    ws_summary.merge_cells("A2:K2")
    date_cell = ws_summary["A2"]
    date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    date_cell.font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    date_cell.alignment = Alignment(horizontal="center")
    ws_summary.row_dimensions[2].height = 18

    # Empty row
    ws_summary.row_dimensions[3].height = 8

    # ---- Summary stats row ----
    total = len(predictions)
    at_risk = sum(1 for p in predictions if p.get("risk_status") == "At Risk")
    normal = total - at_risk
    avg_pct = sum(p.get("current_pct", 0) for p in predictions) / total if total else 0

    for i, (label, val, color) in enumerate([
        ("Total Students", total, COLOR_PRIMARY),
        ("Normal", normal, COLOR_SUCCESS),
        ("At Risk", at_risk, COLOR_DANGER),
        ("Avg Attendance %", f"{avg_pct:.1f}%", "4B5563"),
    ], start=1):
        col_start = (i - 1) * 3 + 1
        col_end = col_start + 1

        # Label cell
        lbl = ws_summary.cell(row=4, column=col_start, value=label)
        lbl.font = Font(name="Calibri", bold=True, size=10, color=COLOR_WHITE)
        lbl.fill = PatternFill("solid", fgColor=COLOR_DARK)
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        lbl.border = thin_border()
        ws_summary.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_end)

        # Value cell
        v = ws_summary.cell(row=5, column=col_start, value=val)
        v.font = Font(name="Calibri", bold=True, size=18, color=color)
        v.fill = PatternFill("solid", fgColor=COLOR_LIGHT)
        v.alignment = Alignment(horizontal="center", vertical="center")
        v.border = thin_border()
        ws_summary.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_end)

    ws_summary.row_dimensions[4].height = 22
    ws_summary.row_dimensions[5].height = 36

    # Empty row
    ws_summary.row_dimensions[6].height = 10

    # ---- Detailed headers ----
    headers = ["Roll No", "Name", "Course", "Semester", "Present", "Absent", "Late",
               "Total Classes", "Current %", "Predicted %", "Risk Status", "Suggestion"]
    col_widths = [12, 22, 16, 10, 10, 10, 8, 14, 12, 14, 12, 50]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws_summary.cell(row=7, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = width
    ws_summary.row_dimensions[7].height = 24

    # ---- Data rows ----
    sorted_preds = sorted(predictions, key=lambda x: x.get("roll_no", ""))
    for row_idx, p in enumerate(sorted_preds, start=8):
        current_pct = p.get("current_pct", 0)
        risk = p.get("risk_status", "")

        if current_pct >= 75:
            att_status = "Good"
            att_color = "D1FAE5"
        elif current_pct >= 60:
            att_status = "Warning"
            att_color = "FEF3C7"
        else:
            att_status = "Critical"
            att_color = "FEE2E2"

        row_data = [
            p.get("roll_no", ""),
            p.get("name", ""),
            p.get("course", ""),
            p.get("semester", ""),
            p.get("present", 0),
            p.get("absent", 0),
            p.get("late", 0),
            p.get("total", 0),
            f"{current_pct:.1f}%",
            f"{p.get('predicted_pct', 0):.1f}%",
            risk,
            p.get("suggestion", ""),
        ]

        bg = COLOR_WHITE if row_idx % 2 == 0 else "F9FAFB"

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(col_idx == 12))
            cell.border = thin_border()

            # Color by attendance status
            if col_idx == 9:
                cell.fill = PatternFill("solid", fgColor=att_color)
                cell.font = Font(name="Calibri", size=9, bold=True,
                                 color=COLOR_SUCCESS if current_pct >= 75 else (COLOR_WARNING if current_pct >= 60 else COLOR_DANGER))
            elif col_idx == 11:
                risk_color = "D1FAE5" if risk == "Normal" else "FEE2E2"
                cell.fill = PatternFill("solid", fgColor=risk_color)
                cell.font = Font(name="Calibri", size=9, bold=True,
                                 color=COLOR_SUCCESS if risk == "Normal" else COLOR_DANGER)
            else:
                cell.fill = PatternFill("solid", fgColor=bg)

        ws_summary.row_dimensions[row_idx].height = 18

    # Freeze header rows
    ws_summary.freeze_panes = "A8"

    # ===== Sheet 2: Raw Attendance =====
    if attendance:
        ws_att = wb.create_sheet("Raw Attendance")
        att_headers = ["Student ID", "Date", "Status"]
        for col_idx, header in enumerate(att_headers, start=1):
            cell = ws_att.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color=COLOR_WHITE)
            cell.fill = PatternFill("solid", fgColor=COLOR_DARK)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, rec in enumerate(attendance, start=2):
            ws_att.cell(row=row_idx, column=1, value=rec.get("student_id", ""))
            ws_att.cell(row=row_idx, column=2, value=rec.get("date", ""))
            status_cell = ws_att.cell(row=row_idx, column=3, value=rec.get("status", ""))
            if rec.get("status") == "Present":
                status_cell.font = Font(color=COLOR_SUCCESS, bold=True)
            elif rec.get("status") == "Absent":
                status_cell.font = Font(color=COLOR_DANGER, bold=True)
            else:
                status_cell.font = Font(color=COLOR_WARNING, bold=True)

        for col in ["A", "B", "C"]:
            ws_att.column_dimensions[col].width = 36 if col == "A" else 14

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
